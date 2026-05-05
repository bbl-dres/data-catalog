#!/usr/bin/env python3
"""
generate-ibpdi-import.py

Convert the IBPDI Real Estate CDM source workbook (one flat sheet,
~2000 rows of entity/attribute/relationship triples) into an XLSX in
the BBL Canvas import format.

Source columns expected (header row 1):
    Data Cluster | Entity name | Attribute Name | Attribute Description |
    Type | Precision | Scale | Min length | Max length | Enum |
    Min value | Max value | Unit | Required | Primary key |
    Relationship type | Relationship name | Cardinality | Examples

Output sheets (canonical names from prototype-canvas/js/xlsx_io.js):
    distribution    one row per entity     headers: id, label, type, system, schema, tags, x, y
    attribute       one row per source row  headers: node_id, name, type, key,
                                                    set_id, source_structure
    edge            one row per FK pair    headers: id, from, to, label
    pset            empty (registry kept clean)
    system          empty (derived from distribution.system on import)

Usage:
    python scripts/generate-ibpdi-import.py [SOURCE_XLSX] [OUTPUT_XLSX]

Defaults:
    SOURCE_XLSX = C:/Users/DavidRasner/Downloads/202404_IBPDI_Real_Estate_CDM.xlsx
    OUTPUT_XLSX = C:/Users/DavidRasner/Downloads/ibpdi_for_bbl_canvas.xlsx
"""

import math
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


DEFAULT_SOURCE = Path(r"C:/Users/DavidRasner/Downloads/202404_IBPDI_Real_Estate_CDM.xlsx")
DEFAULT_OUTPUT = Path(r"C:/Users/DavidRasner/Downloads/ibpdi_for_bbl_canvas.xlsx")

# Layout: per-cluster grid, clusters stacked vertically. Numbers picked so
# 12-15 columns of nodes fit comfortably; resulting canvas is wide but the
# minimap + zoom-extent make it navigable.
NODE_W = 380
NODE_H = 200
COL_GAP = 60
ROW_GAP = 80
COLS_PER_CLUSTER = 14
CLUSTER_GAP = 280   # vertical gap between clusters

# Cluster order mirrors business-significance: Digital Twin first (largest,
# most foundational), then governance/management clusters, then UX last.
CLUSTER_ORDER = [
    "DigitalTwin",
    "PortfolioAndAssetManagement",
    "PropertyManagement",
    "Financials",
    "OrganisationalManagement",
    "EnergyAndResources",
    "UserAndCustomerExperience",
]


def to_snake(name: str) -> str:
    """CamelCase / PascalCase -> snake_case. Keeps acronyms readable
    (e.g. 'HVACSystem' -> 'hvac_system'). Numerics passed through."""
    if not name:
        return ""
    s1 = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)
    s2 = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s1)
    return s2.lower()


def parse_fk(rel_name: str):
    """Parse 'fk A.x to B.y' -> ('A', 'x', 'B', 'y') or None on malformed."""
    if not rel_name:
        return None
    m = re.match(r"^\s*fk\s+([\w]+)\.([\w]+)\s+to\s+([\w]+)\.([\w]+)\s*$", rel_name, re.IGNORECASE)
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3), m.group(4)


def read_source(path: Path):
    """Read the IBPDI source. Yields one dict per data row."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "IBPDI_Real_Estate_CDM" in wb.sheetnames:
        ws = wb["IBPDI_Real_Estate_CDM"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Header row -> column index lookup so we don't depend on positional drift.
    header = [(c.value or "").strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    needed = ["Data Cluster", "Entity name", "Attribute Name", "Type",
              "Primary key", "Relationship type", "Relationship name", "Cardinality"]
    missing = [n for n in needed if n not in idx]
    if missing:
        raise ValueError(f"Source workbook missing expected columns: {missing}")

    def cell(row, name):
        v = row[idx[name]]
        return "" if v is None else str(v).strip()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Entity name"]]:
            continue
        yield {
            "cluster":      cell(row, "Data Cluster"),
            "entity":       cell(row, "Entity name"),
            "attribute":    cell(row, "Attribute Name"),
            "type":         cell(row, "Type"),
            "primary_key":  cell(row, "Primary key").upper() == "TRUE",
            "rel_type":     cell(row, "Relationship type"),
            "rel_name":     cell(row, "Relationship name"),
            "cardinality":  cell(row, "Cardinality"),
        }


JUNCTION_PREFIXES = ("Role",)
"""Wrapper words that turn an A+B junction name into a still-junction name —
`RoleComponentContact` is a Component↔Contact role-assignment table, same
visual clutter as `ComponentSpace` for graph-view purposes."""


def detect_junctions(entity_attrs, entity_fk_targets):
    """Identify pure many-to-many relation tables that exist only to bridge
    two real entities (e.g. `ComponentSpace` between `Component` and
    `Space`, or `RoleComponentContact` between `Component` and `Contact`).
    They add no business meaning of their own and clutter the graph view.

    Heuristic — entity is a junction if ALL of:
      1. ≥ 2 FK columns to ≥ 2 distinct target entities (must actually be a
         relation, not a regular table that happens to have a couple of FKs)
      2. Name == A+B (in either order) for some pair of target entities,
         optionally with a `Role` prefix — this is the strong signal that
         it's purely a bridge:
            ComponentSpace        ↔ {Component, Space}
            RoleComponentContact  ↔ {Component, Contact}
      3. Total attribute count is small (≤ 10) — pure junctions carry only
         their FK columns + maybe an audit field; role tables carry a few
         more (role enum, validity dates) but still nothing that warrants
         their own lifecycle. This rules out regular entities that happen
         to share a name pattern (e.g. `UserAccount` if both `User` and
         `Account` exist as proper entities with rich attribute sets)

    Returns a set of entity names judged to be junctions.
    """
    junctions = set()
    for entity, targets in entity_fk_targets.items():
        unique_targets = list(set(targets))
        if len(unique_targets) < 2:
            continue
        if len(entity_attrs.get(entity, [])) > 10:
            continue

        # Build the list of name forms to test against A+B concat. For
        # `RoleComponentContact` we also test `ComponentContact` so the
        # Role-prefix wrapper doesn't hide the junction shape.
        candidate_names = [entity]
        for prefix in JUNCTION_PREFIXES:
            if entity.startswith(prefix) and len(entity) > len(prefix):
                candidate_names.append(entity[len(prefix):])

        is_concat = False
        for cn in candidate_names:
            for a in unique_targets:
                for b in unique_targets:
                    if a != b and cn == a + b:
                        is_concat = True
                        break
                if is_concat:
                    break
            if is_concat:
                break
        if is_concat:
            junctions.add(entity)
    return junctions


def build(rows):
    """Walk source rows once, produce distribution/attribute/edge lists.
    Returns (distributions, attributes, edges)."""
    # entity_to_cluster: entity name -> cluster (first seen wins; entities
    # appear in many rows but always belong to one cluster).
    entity_to_cluster = {}
    entity_attrs = defaultdict(list)         # entity_name -> list of attr dicts
    entity_fk_targets = defaultdict(list)    # entity_name -> list of target entity names (for junction detection)
    fk_pairs = defaultdict(list)             # (from_slug, to_slug) -> list of attr names
    skipped_fks = 0

    for r in rows:
        cluster, entity, attr_name = r["cluster"], r["entity"], r["attribute"]
        if not entity or not attr_name:
            continue
        entity_to_cluster.setdefault(entity, cluster)

        # Attribute key role.
        key = ""
        if r["primary_key"]:
            key = "PK"
        elif r["rel_type"].lower() == "foreign key":
            key = "FK"

        entity_attrs[entity].append({
            "name": attr_name,
            "type": r["type"] or "",
            "key":  key,
        })

        # Foreign-key edge.
        if r["rel_type"].lower() == "foreign key":
            parsed = parse_fk(r["rel_name"])
            if not parsed:
                skipped_fks += 1
                continue
            src_ent, src_attr, dst_ent, dst_attr = parsed
            entity_fk_targets[src_ent].append(dst_ent)
            from_slug = to_snake(src_ent)
            to_slug   = to_snake(dst_ent)
            if from_slug == to_slug:
                # Self-loop — the BBL edge model rejects these; skip cleanly.
                continue
            fk_pairs[(from_slug, to_slug)].append(src_attr)

    if skipped_fks:
        print(f"  (warn) {skipped_fks} FK rows had unparseable relationship names — skipped")

    # Detect + report junction tables before any output is built.
    junctions = detect_junctions(entity_attrs, entity_fk_targets)
    junction_slugs = {to_snake(j) for j in junctions}
    if junctions:
        print(f"  Detected {len(junctions)} junction tables — dropping + bridging edges:")
        for j in sorted(junctions):
            tgt_list = sorted(set(entity_fk_targets[j]))
            print(f"    - {j} (FKs -> {', '.join(tgt_list)})")

    # Distributions, with per-cluster grid layout. Junctions are skipped
    # so they never appear in the canvas.
    distributions = []
    by_cluster = defaultdict(list)
    for ent, clu in entity_to_cluster.items():
        if ent in junctions:
            continue
        by_cluster[clu].append(ent)
    for clu in by_cluster:
        by_cluster[clu].sort()  # alphabetical within cluster, stable across runs

    cluster_y = 0.0
    for clu in CLUSTER_ORDER + [c for c in sorted(by_cluster) if c not in CLUSTER_ORDER]:
        ents = by_cluster.get(clu, [])
        if not ents:
            continue
        for i, ent in enumerate(ents):
            col = i % COLS_PER_CLUSTER
            row = i // COLS_PER_CLUSTER
            x = col * (NODE_W + COL_GAP)
            y = cluster_y + row * (NODE_H + ROW_GAP)
            distributions.append({
                "id":     to_snake(ent),
                "label":  ent,
                "type":   "table",
                "system": clu,
                "schema": "",
                "tags":   "",
                "x":      x,
                "y":      y,
            })
        rows_needed = math.ceil(len(ents) / COLS_PER_CLUSTER)
        cluster_y += rows_needed * (NODE_H + ROW_GAP) + CLUSTER_GAP

    # Attribute rows, in source order per entity. Skip attributes belonging
    # to junctions — those entities are gone from distributions.
    attributes = []
    for ent, attrs in entity_attrs.items():
        if ent in junctions:
            continue
        node_id = to_snake(ent)
        for a in attrs:
            attributes.append({
                "node_id":          node_id,
                "name":             a["name"],
                "type":             a["type"],
                "key":              a["key"],
                "set_id":           "",
                "source_structure": "",
            })

    # Edges — one per (from, to) pair with combined labels. Drop any edge
    # touching a junction (its endpoints are gone), then synthesise bridge
    # edges between every pair of entities the junction connected.
    valid_ids = {d["id"] for d in distributions}
    edges = []
    edge_pairs = set()  # (from, to) -> dedupe across original + synthetic edges

    for (frm, to), attr_names in sorted(fk_pairs.items()):
        if frm in junction_slugs or to in junction_slugs:
            continue  # bridged below
        if frm not in valid_ids or to not in valid_ids:
            print(f"  (warn) edge dropped — endpoint missing: {frm} -> {to}")
            continue
        unique_attrs = sorted(set(attr_names))
        edges.append({
            "id":    f"e_{frm}__{to}",
            "from":  frm,
            "to":    to,
            "label": ", ".join(unique_attrs),
        })
        edge_pairs.add((frm, to))

    # Synthesise bridge edges. For a junction with FKs to {A, B, …} we emit
    # one edge per unordered pair of targets; the junction's name is the
    # edge label (so the visual still tells you "A is associated with B
    # via JunctionTable"). Sorted endpoints give a stable canonical form
    # — unordered relationships shouldn't depend on FK declaration order.
    bridged = 0
    for j in sorted(junctions):
        targets = sorted(set(entity_fk_targets[j]))
        target_slugs = [to_snake(t) for t in targets]
        for i in range(len(target_slugs)):
            for k in range(i + 1, len(target_slugs)):
                a, b = sorted([target_slugs[i], target_slugs[k]])
                if a not in valid_ids or b not in valid_ids:
                    continue
                if (a, b) in edge_pairs or (b, a) in edge_pairs:
                    continue  # original FK already wired these two — don't duplicate
                edges.append({
                    "id":    f"e_{a}__{b}__via_{to_snake(j)}",
                    "from":  a,
                    "to":    b,
                    "label": j,
                })
                edge_pairs.add((a, b))
                bridged += 1

    if junctions:
        print(f"  Bridged {bridged} synthetic edges across {len(junctions)} junctions")

    return distributions, attributes, edges


def write_workbook(out_path: Path, distributions, attributes, edges):
    """Emit the BBL-Canvas-shape import workbook."""
    wb = openpyxl.Workbook()
    # openpyxl creates a default empty sheet; rename + use it for the first.
    ws = wb.active
    ws.title = "distribution"

    # --- distribution ---
    dist_headers = ["id", "label", "type", "system", "schema", "tags", "x", "y"]
    ws.append(dist_headers)
    for d in distributions:
        ws.append([d.get(h, "") for h in dist_headers])

    # --- attribute ---
    ws_a = wb.create_sheet("attribute")
    attr_headers = ["node_id", "name", "type", "key", "set_id", "source_structure"]
    ws_a.append(attr_headers)
    for a in attributes:
        ws_a.append([a.get(h, "") for h in attr_headers])

    # --- edge ---
    ws_e = wb.create_sheet("edge")
    edge_headers = ["id", "from", "to", "label"]
    ws_e.append(edge_headers)
    for e in edges:
        ws_e.append([e.get(h, "") for h in edge_headers])

    # --- pset (empty, but the sheet must exist for the importer to find it
    # and not fall back to State.getSets()) ---
    ws_p = wb.create_sheet("pset")
    ws_p.append(["id", "label", "description", "lineage"])

    # --- system (auxiliary; importer derives from distribution.system) ---
    ws_s = wb.create_sheet("system")
    ws_s.append(["name", "nodes", "tables", "apis", "files", "valuelists",
                 "sets", "attributes", "tags"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not src.exists():
        print(f"Source file not found: {src}")
        sys.exit(1)

    print(f"Reading {src}")
    rows = list(read_source(src))
    print(f"  {len(rows)} attribute rows")

    distributions, attributes, edges = build(rows)
    print(f"Built:")
    print(f"  {len(distributions)} distributions")
    print(f"  {len(attributes)} attributes")
    print(f"  {len(edges)} edges (from FK pairs, deduped)")

    write_workbook(out, distributions, attributes, edges)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
