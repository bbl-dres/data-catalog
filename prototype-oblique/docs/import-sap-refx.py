"""Import SAP API documentation and the reviewed Innovator diagram transcription.

Requires beautifulsoup4; openpyxl is needed only for the earlier workbook export.
Model identifiers and semantic API candidates never imply physical SAP columns.
"""
import argparse
import hashlib
import json
from pathlib import Path

from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[1]
IMPORT_ID = 'sap-refx-building-documentation'
API_ID = 'api-sap-building'
MODEL_IMPORT_ID = 'sap-refx-innovator-model'
LEGACY_TABLE_IDS = {'t-we', 't-geb-sap', 't-mo', 't-mv'}
DOCUMENTATION = 'https://confluence.bit.admin.ch/pages/viewpage.action?pageId=1105159761'


def clean(value):
    return ' '.join(str(value or '').split())


def text(node):
    return clean(node.get_text(' ', strip=True)) if node else ''


def table_rows(table):
    return [[text(cell) for cell in row.find_all(['th', 'td'], recursive=False)]
            for row in table.find_all('tr')]


def extract_page(path):
    soup = BeautifulSoup(path.read_bytes(), 'html.parser')
    main = soup.find(id='main-content')
    if not main:
        raise ValueError('Missing Confluence page content')
    metadata = {}
    prefix = 'apiOverview.intro.apiMetadata.'
    for node in main.find_all(id=True):
        key = node['id']
        if key.startswith(prefix) and not key.endswith('.title'):
            metadata[key[len(prefix):]] = text(node)
    if metadata.get('apiName') != 'ZAPI_X4AI_BAPI_RE_BU_GET_DET':
        raise ValueError('Unexpected SAP API; review this importer before using another page')

    entities, attributes = {}, {}
    for table in main.find_all('table'):
        heading = table.find_previous(['h1', 'h2', 'h3', 'h4', 'h5'])
        rows = table_rows(table)
        if rows[0] == ['Entität', 'Service Knoten', 'Beschreibung'] and len(rows) > 2:
            entities = {row[1]: {'name': row[0], 'description': row[2]} for row in rows[1:]}
        elif text(heading).startswith('Attribute '):
            name = text(heading).removeprefix('Attribute ')
            fields, includes = [], []
            for number, row in enumerate(rows[1:], 2):
                if len(row) != 4:
                    raise ValueError(f'{name} row {number}: expected four cells')
                technical_name, description, data_type, length = row
                if technical_name == '.INCLUDE':
                    includes.append({'row': number, 'description': description})
                    continue
                fields.append({'technicalName': technical_name, 'description': description,
                               'reportedDataType': data_type or None,
                               'reportedLength': int(length) if length else None, 'row': number})
            if len({field['technicalName'] for field in fields}) != len(fields):
                raise ValueError(f'{name}: duplicate field names')
            attributes[name] = {'fields': fields, 'includes': includes,
                                'sourceUrl': DOCUMENTATION + '#' + heading['id']}
    if len(entities) != 25 or set(entities) != set(attributes):
        raise ValueError('Service structure coverage changed; review before importing')
    structures = [{'technicalName': key, **entity, **attributes[key]} for key, entity in entities.items()]
    servers = [{'url': option['value'], 'description': text(option).removeprefix(option['value']).strip(' -')}
               for option in main.select('#swagger-ui select option')]
    if len(servers) != 3 or not all(server['url'].startswith('https://') for server in servers):
        raise ValueError('Expected three documented HTTPS servers')
    return metadata, structures, servers


def inspect_model(path):
    if path.suffix.lower() == '.png':
        model_path = ROOT / 'docs/sap-refx-model.json'
        matches_path = ROOT / 'docs/sap-refx-field-matches.json'
        model = json.loads(model_path.read_text(encoding='utf-8'))
        if hashlib.sha256(path.read_bytes()).hexdigest() != model['sourceSha256']:
            raise ValueError('Diagram changed; review the transcription before importing')
        model['status'] = 'transcribed'
        model['fieldMatches'] = json.loads(matches_path.read_text(encoding='utf-8'))
        model['transcriptionSha256'] = hashlib.sha256(model_path.read_bytes()).hexdigest()
        model['fieldMatchesSha256'] = hashlib.sha256(matches_path.read_bytes()).hexdigest()
        return model
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for sheet in workbook:
            rows = [[clean(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            rows = [row for row in rows if any(row)]
            sheets.append({'name': sheet.title, 'headers': rows[0] if rows else [],
                           'recordCount': max(0, len(rows) - 1)})
        # This metadata-only export cannot support a field-level match.
        metadata_only = (len(sheets) == 1 and sheets[0]['name'] == 'Class Diagram (Application)'
                         and sheets[0]['recordCount'] == 1)
        if not metadata_only:
            raise ValueError('The workbook now contains different content; implement the model mapping before importing')
        return {'status': 'missing-attributes', 'sheets': sheets,
                'note': 'The workbook contains one diagram metadata record and no classes, attributes or embedded diagram.'}
    finally:
        workbook.close()


def write_json(path, value):
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')


def preserve_annotations(record, previous):
    """Keep catalog comments and links when refreshing an owned source record."""
    previous = previous or {}
    for key in ('comment', 'informationUrls'):
        if key in previous:
            record[key] = previous[key]
    fields = {field.get('identifier', field['technicalName']): field for field in previous.get('fields', [])}
    for field in record.get('fields', []):
        old = fields.get(field.get('identifier', field['technicalName']), {})
        if 'comment' in old:
            field['comment'] = old['comment']


def apply_table_context(tables, decisions, captured):
    """Separate documented definitions from catalog review notes and field provenance."""
    path = ROOT / 'docs/sap-refx-definitions.json'
    snapshot = json.loads(path.read_text(encoding='utf-8'))
    by_id = {table['identifier']: table for table in tables}
    ids = [entry['tableId'] for entry in snapshot['entries']]
    if len(ids) != len(set(ids)) or set(ids) - by_id.keys():
        raise ValueError('Definition references a duplicate or unknown table')
    if set(decisions.get('tableComments', {})) - by_id.keys():
        raise ValueError('Comment references an unknown table')
    for entry in snapshot['entries']:
        if entry['kind'] not in {'source-excerpt', 'source-summary'}:
            raise ValueError('Unknown definition source kind')
        table = by_id[entry['tableId']]
        table.update(description=entry['description'], modified=captured,
                     descriptionSource={'title': entry['title'], 'url': entry['sourceUrl'], 'kind': entry['kind'],
                                        'reviewed': snapshot['reviewed'], 'version': snapshot['sourceVersion']})
        table['informationUrls'] = list(dict.fromkeys([*table.get('informationUrls', []), entry['sourceUrl']]))
    for table in tables:
        if table['identifier'] in decisions.get('tableComments', {}):
            table.setdefault('comment', decisions['tableComments'][table['identifier']])
        # First imports and refreshes must serialize curated metadata in the same order.
        for key in ('descriptionSource', 'informationUrls', 'comment'):
            if key in table:
                table[key] = table.pop(key)
    return {'snapshot': snapshot, 'sha256': hashlib.sha256(path.read_bytes()).hexdigest()}


def import_data_sources(tables, selections, curation_hash, captured):
    """Import reviewed extraction inventories without inferring physical column contracts."""
    path = ROOT / 'docs/sap-refx-data-sources.json'
    snapshot = json.loads(path.read_text(encoding='utf-8'))
    sources = {source['identifier']: source for source in snapshot['sources']}
    if len(sources) != len(snapshot['sources']):
        raise ValueError('Duplicate DataSource identifier')
    if len({item['tableId'] for item in selections}) != len(selections):
        raise ValueError('Duplicate DataSource table identifier')
    domains = {item['identifier'] for item in json.loads((ROOT / 'data/domains.json').read_text(encoding='utf-8'))}
    objects = {item['identifier'] for item in json.loads((ROOT / 'data/objects.json').read_text(encoding='utf-8'))}
    snapshot_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    for item in selections:
        source = sources[item['dataSource']]
        if item['domain'] not in domains or (item.get('realizes') and item['realizes'] not in objects):
            raise ValueError('DataSource entry references an unknown domain or business object')
        ids = [field['technicalName'] for field in source['fields']]
        if len(ids) != len(set(ids)) or any(not isinstance(name, str) or not name.strip() for name in ids):
            raise ValueError('Invalid or duplicate extraction field identifier')
        if any(field['sourceTable'] != source['technicalName'] for field in source['fields']):
            raise ValueError('Extraction combines source tables; review its catalog scope before importing')
        previous = next((table for table in tables if table['identifier'] == item['tableId']), None)
        if previous and previous.get('provenance', {}).get('importId') != MODEL_IMPORT_ID:
            raise ValueError('DataSource table identifier is owned by another record')
        record = {
            'identifier': item['tableId'], 'name': item['name'], 'labels': {'de': item['name']},
            'technicalName': source['technicalName'], 'technicalNameKind': 'physical-table',
            'technicalNameSource': source['sourceUrl'], 'description': source['description'],
            'status': 'Entwurf', 'created': previous['created'] if previous else captured, 'modified': captured,
            'responsibleOrg': 'Bundesamt für Bauten und Logistik BBL', 'system': 'sap',
            'domain': item['domain'],
            'source': 'SAP Help Portal', 'sourceUrl': source['sourceUrl'],
            'sourceDetail': f"{source['title']} · {source['identifier']} · Dokumentierte Extraktionsfelder; keine vollständige physische Tabellendefinition",
            'fieldScope': 'datasource-projection', 'dataSource': source['identifier'], 'synced': captured,
            'fields': [{'technicalName': field['technicalName'], 'technicalNameKind': 'datasource-field',
                        'labels': {'de': field['label']}, 'description': field['label'], 'keyRole': None,
                        'sourceUrl': source['sourceUrl'],
                        'catalogMetadata': {'dataSource': source['identifier'], 'sourceTable': field['sourceTable']}}
                       for field in source['fields']],
            'provenance': {'importId': MODEL_IMPORT_ID, 'captured': captured,
                           'sourceCaptured': snapshot['captured'], 'sourceSnapshot': path.name,
                           'sourceSnapshotSha256': snapshot_hash, 'curationSha256': curation_hash},
        }
        if item.get('realizes'):
            record['realizes'] = item['realizes']
        preserve_annotations(record, previous)
        if previous:
            tables[tables.index(previous)] = record
        else:
            tables.append(record)
    return {'snapshot': snapshot, 'sha256': snapshot_hash,
            'tableIds': [item['tableId'] for item in selections]}


def curate_catalog(tables, model, structures, provenance, captured):
    """Apply catalog decisions after reconciling the unchanged source transcription."""
    path = ROOT / 'docs/sap-refx-catalog-curation.json'
    decisions = json.loads(path.read_text(encoding='utf-8'))
    classes = {item['name']: item for item in model['classes']}
    if set(decisions['excludeClasses']) - classes.keys():
        raise ValueError('Catalog exclusion references an unknown source class')
    excluded = [item for item in classes.values()
                if item['view'] in decisions['excludeViews'] or item['name'] in decisions['excludeClasses']]
    retired = {item['tableId'] for item in excluded}
    changes = [{'sourceClass': item['name'], 'tableId': item['tableId'], 'action': 'excluded',
                'reason': decisions['reason']} for item in excluded]
    by_id = {table['identifier']: table for table in tables}
    curation_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    for group in decisions['consolidations']:
        sources = [by_id[classes[name]['tableId']] for name in group['classes']]
        if any(source['identifier'] in retired for source in sources):
            raise ValueError('Consolidation overlaps an excluded or already consolidated class')
        fields, types = [], []
        for source in sources:
            name = source['modelClass']
            type_fields = []
            for field in source['fields']:
                fields.append({**field, 'appliesToObjectTypes': [name],
                               'description': f'Im Quelldiagramm beim Typ {name} aufgeführt; keine Aussage zur Verwendung bei anderen Typen.'})
                type_fields.append(field.get('identifier', field['technicalName']))
            types.append({'name': name, 'businessObject': source.get('realizes'),
                          'fieldIds': type_fields, 'sourceClass': name,
                          'sourceBounds': source['provenance']['diagramBounds'],
                          'sourceAssociations': source['modelAssociations']})
            retired.add(source['identifier'])
            changes.append({'sourceClass': name, 'tableId': source['identifier'], 'action': 'consolidated',
                            'replacement': group['tableId'], 'reason': group['reason']})
        ids = [field.get('identifier', field['technicalName']) for field in fields]
        if len(ids) != len(set(ids)):
            raise ValueError('Type-specific fields need distinct identifiers before consolidation')
        previous = by_id.get(group['tableId'])
        if previous and previous.get('provenance', {}).get('importId') != MODEL_IMPORT_ID:
            raise ValueError('Consolidated table identifier is owned by another record')
        type_provenance = {key: value for key, value in sources[0]['provenance'].items() if key != 'diagramBounds'}
        record = {
            'identifier': group['tableId'], 'name': group['name'], 'technicalName': group['name'],
            'technicalNameKind': 'catalog-name',
            'description': 'Architektonisches Objekt mit den Typen Ebene und Raum. Die Felder aus dem Quelldiagramm sind jeweils einem Typ zugeordnet; sie bilden keine bestätigte gemeinsame physische SAP-Felddefinition.',
            'status': 'Entwurf', 'created': previous['created'] if previous else captured, 'modified': captured,
            'responsibleOrg': 'Bundesamt für Bauten und Logistik BBL',
            'source': 'Innovator; fachliche Katalogkorrektur',
            'sourceDetail': 'SAP RE-FX.png · Ebene und Raum als Objekttypen zusammengeführt · Physische SAP-Namen und gemeinsame Felder nicht bestätigt',
            'synced': captured, 'system': 'sap', 'domain': group['domain'],
            'objectTypes': types, 'fields': fields,
            'provenance': {**type_provenance, 'curationSha256': curation_hash, 'curationAuthority': decisions['authority']},
        }
        if group.get('technicalName'):
            record.update(technicalName=group['technicalName'], technicalNameKind='physical-table',
                          technicalNameSource=group['technicalNameSource'])
        preserve_annotations(record, previous)
        if previous:
            tables[tables.index(previous)] = record
        else:
            tables.append(record)
    remaining = [table for table in tables if table['identifier'] not in retired]
    for table in remaining:
        if table.get('provenance', {}).get('importId') == MODEL_IMPORT_ID:
            if table.get('technicalNameKind') in {'model-class', 'catalog-name'}:
                table.pop('technicalName', None)
                table.pop('technicalNameKind', None)
            table['name'] = decisions.get('classAliases', {}).get(table.get('modelClass'), table['name'])
            table['labels'] = {'de': table['name']}
    for item in decisions.get('apiTables', []):
        replaced = classes[item['replacesClass']]['tableId']
        remaining = [table for table in remaining if table['identifier'] != replaced]
        changes.append({'sourceClass': item['replacesClass'], 'tableId': replaced, 'action': 'replaced-by-api-projection',
                        'replacement': item['tableId'], 'reason': item['reason']})
        structure = next(structure for structure in structures if structure['technicalName'] == item['structure'])
        fields = []
        for field in structure['fields']:
            technical_name = field['technicalName']
            fields.append({'technicalName': technical_name, 'technicalNameKind': 'api-field',
                           'labels': {'de': field['description'].removeprefix('BAPI: ') or technical_name},
                           'description': field['description'], 'keyRole': None,
                           'sourceUrl': structure['sourceUrl'],
                           'apiMappings': [{'api': API_ID, 'structure': structure['technicalName'], 'field': technical_name,
                                            'matchType': 'documented-api-field', 'verification': 'documented-api-field',
                                            'physicalColumnVerified': False, 'sourceUrl': structure['sourceUrl']}],
                           'catalogMetadata': {'apiStructure': structure['technicalName'], 'sourceField': technical_name}})
        previous = by_id.get(item['tableId'])
        if previous and previous.get('provenance', {}).get('importId') != MODEL_IMPORT_ID:
            raise ValueError('API projection identifier is owned by another record')
        record = {
            'identifier': item['tableId'], 'name': item['name'], 'labels': {'de': item['name']},
            'technicalName': item['technicalName'], 'technicalNameKind': 'physical-table',
            'technicalNameSource': item['technicalNameSource'],
            'description': 'Gebäudestammdaten aus SAP RE-FX. Feldnamen und Beschreibungen stammen aus der API-Struktur BUILDING; die Zuordnung zu physischen Spalten der Tabelle VIBDBU ist noch nicht bestätigt.',
            'status': 'Entwurf', 'created': previous['created'] if previous else captured, 'modified': captured,
            'responsibleOrg': 'Bundesamt für Bauten und Logistik BBL', 'system': 'sap',
            'domain': item['domain'], 'realizes': item['realizes'], 'fields': fields,
            'source': 'SAP API Dokumentation Bund', 'sourceUrl': structure['sourceUrl'],
            'sourceDetail': 'Building Master Data – Get Detail · BUILDING · Technische API-Feldnamen; keine verifizierte physische Spaltendefinition',
            'fieldScope': 'api-projection', 'apiStructure': structure['technicalName'], 'synced': captured,
            'provenance': {**provenance, 'importId': MODEL_IMPORT_ID, 'curationSha256': curation_hash,
                           'curationAuthority': decisions['authority']},
        }
        existing = next((table for table in remaining if table['identifier'] == item['tableId']), None)
        preserve_annotations(record, existing)
        if existing:
            remaining[remaining.index(existing)] = record
        else:
            remaining.append(record)
    source_import = import_data_sources(remaining, decisions.get('dataSourceTables', []), curation_hash, captured)
    for table in remaining:
        if table.get('provenance', {}).get('importId') != MODEL_IMPORT_ID:
            continue
        urls = list(dict.fromkeys([*table.get('informationUrls', []), table.get('sourceUrl'), table.get('technicalNameSource')]))
        if any(urls):
            table['informationUrls'] = [url for url in urls if url]
    definition_import = apply_table_context(remaining, decisions, captured)
    active = [table for table in remaining if table.get('provenance', {}).get('importId') == MODEL_IMPORT_ID]
    return remaining, {'decisions': decisions, 'sha256': curation_hash, 'changes': changes,
                       'dataSources': source_import,
                       'definitions': definition_import,
                       'activeTableIds': [table['identifier'] for table in active],
                       'activeTables': len(active), 'activeFields': sum(len(table['fields']) for table in active)}


def retire_references(curation, previous):
    """Remove retired catalog links while retaining their original source evidence."""
    retired = {item['tableId'] for item in curation['changes']}
    archive = previous.get('catalogCuration', {}).get('retiredReferences', {'productSources': [], 'history': []})
    products_path, history_path = ROOT / 'data/products.json', ROOT / 'data/changelog.json'
    products = json.loads(products_path.read_text(encoding='utf-8'))
    for product in products:
        removed = [table_id for table_id in product.get('sourcedFrom', []) if table_id in retired]
        if removed:
            archive['productSources'].append({'product': product['identifier'], 'tables': removed})
            product['sourcedFrom'] = [table_id for table_id in product['sourcedFrom'] if table_id not in retired]
    history = json.loads(history_path.read_text(encoding='utf-8'))
    retired_entities = {'tables:' + table_id for table_id in retired}
    archive['history'].extend(item for item in history if item['entity'] in retired_entities)
    curation['retiredReferences'] = archive
    return [(products_path, products), (history_path, [item for item in history if item['entity'] not in retired_entities])]


def reconcile_model(model, structures, provenance, captured):
    tables = json.loads((ROOT / 'data/tables.json').read_text(encoding='utf-8'))
    domains = {item['identifier'] for item in json.loads((ROOT / 'data/domains.json').read_text(encoding='utf-8'))}
    objects = {item['identifier'] for item in json.loads((ROOT / 'data/objects.json').read_text(encoding='utf-8'))}
    source_fields = {f"{structure['technicalName']}.{field['technicalName']}": (structure, field)
                     for structure in structures for field in structure['fields']}
    classes = model['classes']
    if len({item['name'] for item in classes}) != len(classes) or len({item['tableId'] for item in classes}) != len(classes):
        raise ValueError('Duplicate model class or table identifier')
    matches = model['fieldMatches']['classes']
    if set(matches) - {item['name'] for item in classes}:
        raise ValueError('API match references an unknown model class')
    candidates, unmatched, used_targets, removals = [], [], set(), []
    for item in classes:
        names = item['attributes']
        if len(names) != len(set(names)) or not all(isinstance(name, str) and name.strip() for name in names):
            raise ValueError(f"{item['name']}: invalid or duplicate attribute names")
        if item['domain'] not in domains or (item.get('realizes') and item['realizes'] not in objects):
            raise ValueError(f"{item['name']}: unknown catalog reference")
        class_matches = matches.get(item['name'], {})
        if set(class_matches) - set(names):
            raise ValueError(f"{item['name']}: API match references an unknown attribute")
        fields = []
        for name in names:
            field = {'technicalName': name, 'labels': {'de': name}, 'description': '', 'keyRole': None,
                     'technicalNameKind': 'model-attribute'}
            field_id = item.get('legacyFieldIds', {}).get(name, name)
            if field_id != name:
                field['identifier'] = field_id
            model_type = item.get('modelTypes', {}).get(name)
            if model_type:
                field['dataType'] = model_type
                field['dataTypeKind'] = 'model-type'
            field['catalogMetadata'] = {'modelClass': item['name'], 'modelAttribute': name,
                                        'sourceFile': model['sourceFile']}
            targets = class_matches.get(name, [])
            if len(targets) != len(set(targets)) or any(target not in source_fields for target in targets):
                raise ValueError(f"{item['name']}.{name}: invalid API target")
            if targets:
                field['apiMappings'] = [
                    {'api': API_ID, 'structure': source_fields[target][0]['technicalName'],
                     'field': source_fields[target][1]['technicalName'], 'matchType': 'semantic-candidate',
                     'verification': 'unverified', 'sourceUrl': source_fields[target][0]['sourceUrl']}
                    for target in targets]
                if len(targets) == 1:
                    field['description'] = source_fields[targets[0]][1]['description']
                    field['descriptionSource'] = 'api-semantic-candidate'
                candidates.append({'table': item['tableId'], 'modelClass': item['name'],
                                   'fieldId': field_id, 'modelAttribute': name, 'targets': targets,
                                   'status': 'ambiguous' if len(targets) > 1 else 'candidate'})
                used_targets.update(targets)
            else:
                unmatched.append({'table': item['tableId'], 'modelClass': item['name'],
                                  'fieldId': field_id, 'modelAttribute': name})
            fields.append(field)
        ids = [field.get('identifier', field['technicalName']) for field in fields]
        if len(ids) != len(set(ids)):
            raise ValueError(f"{item['name']}: duplicate field identifiers after bookmark migration")
        view = 'architektonischen Sicht' if item['view'] == 'architecture' else 'Nutzungssicht'
        description = f"Reduzierte Anwendungsklasse {item['name']} der {view} im SAP RE-FX-Systemmodell des BBL."
        if not fields:
            description += ' Im Diagramm sind keine skalaren Attribute für diese Klasse angegeben.'
        if class_matches:
            description += ' Feldbeschreibungen aus der API beruhen auf fachlichen Zuordnungsvorschlägen; die technische Zuordnung ist noch zu bestätigen.'
        record = {
            'identifier': item['tableId'], 'name': item['title'], 'technicalName': item['name'],
            'technicalNameKind': 'model-class', 'description': description,
            'status': 'Entwurf', 'created': captured, 'modified': captured,
            'responsibleOrg': 'Bundesamt für Bauten und Logistik BBL',
            'source': 'Innovator', 'sourceDetail': 'SAP RE-FX.png · Reduziertes Systemmodell · Modellnamen; physische SAP-Spalten und Datentypen nicht dokumentiert',
            'synced': captured, 'system': 'sap', 'domain': item['domain'],
            'modelClass': item['name'], 'modelView': item['view'], 'modelAbstract': item.get('abstract', False),
            'modelAssociations': item['associations'], 'fields': fields,
            'provenance': {**provenance, 'importId': MODEL_IMPORT_ID, 'diagramBounds': item['bounds'],
                           'transcriptionSha256': model['transcriptionSha256'],
                           'fieldMatchesSha256': model['fieldMatchesSha256']},
        }
        if item.get('realizes'):
            record['realizes'] = item['realizes']
        previous = next((table for table in tables if table['identifier'] == item['tableId']), None)
        if previous:
            owned = previous.get('provenance', {}).get('importId') == MODEL_IMPORT_ID
            if not owned and item['tableId'] not in LEGACY_TABLE_IDS:
                raise ValueError(f"{item['tableId']}: table identifier is owned by another import")
            if owned:
                record['created'] = previous['created']
            preserve_annotations(record, previous)
            replacement = previous.get('prototypeReplacement')
            if not owned:
                replacement = {'technicalName': previous.get('technicalName'),
                               'removedFieldIds': [field.get('identifier', field['technicalName'])
                                                   for field in previous['fields']
                                                   if field.get('identifier', field['technicalName']) not in ids]}
            if replacement:
                record['prototypeReplacement'] = replacement
                removals.append({'table': item['tableId'], **replacement})
            tables[tables.index(previous)] = record
        else:
            tables.append(record)
    api_only = [{'structure': source_fields[target][0]['technicalName'],
                 'field': source_fields[target][1]['technicalName']} for target in source_fields if target not in used_targets]
    summary = {'classes': len(classes), 'modelFields': len(candidates) + len(unmatched),
               'emptyClasses': sum(not item['attributes'] for item in classes),
               'candidateModelFields': len(candidates),
               'ambiguousModelFields': sum(item['status'] == 'ambiguous' for item in candidates),
               'matchedApiFields': len(used_targets), 'unmatchedModelFields': len(unmatched),
               'unmatchedApiFields': len(api_only), 'verifiedPhysicalTableMappings': 0}
    return tables, {'summary': summary, 'candidates': candidates, 'unmatchedModelFields': unmatched,
                    'unmatchedApiFields': api_only, 'prototypeReplacements': removals,
                    'notes': model['fieldMatches']['notes']}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--api-page', type=Path, required=True)
    parser.add_argument('--model', type=Path, required=True, help='Reviewed diagram PNG or the metadata-only workbook')
    parser.add_argument('--captured', required=True, help='Catalog import date, YYYY-MM-DD')
    args = parser.parse_args()
    from datetime import date
    date.fromisoformat(args.captured)
    metadata, structures, servers = extract_page(args.api_page)
    model = inspect_model(args.model)
    has_diagram = model['status'] == 'transcribed'
    provenance = {'importId': IMPORT_ID, 'captured': args.captured,
                  'files': {path.name: hashlib.sha256(path.read_bytes()).hexdigest()
                            for path in (args.api_page, args.model)}}
    issues = [
        {'id': 'missing-physical-schema', 'detail': 'The PNG provides model classes and attributes, but no verified physical column names, database types or constraints.'}
        if has_diagram else {'id': 'missing-model-attributes', 'detail': model['note']},
        {'id': 'building-types-misaligned', 'structures': ['BUILDING'],
         'detail': 'Examples: CREATION_USER is DATS(8), CREATION_DATE is TIMS(6), BUILDING_TEXT is NUMC(2); FUNCTION has no type or length. Do not use the reported types as a verified schema.'},
        {'id': 'property-tax-copy-errors', 'structures': ['PROP_TAX'],
         'detail': 'Multiple descriptions and types repeat OPTION_RATE or unrelated fields. MEAS_TYPE_BUILDING_AREA and DATA_REPORTED have no descriptions or types.'},
        {'id': 'request-response-conflict',
         'detail': 'Prose expects BUILDING_ID; BUILDING lists COMP_CODE, BUSINESS_ENTITY and BUILDING. Examples use DESCRIPTION, STATUS, POSTAL_CODE and CITY, which do not match the documented BUILDING fields.'},
        {'id': 'transport-endpoint-conflict',
         'detail': 'Documentation identifies SOAP but rendered OpenAPI servers use an /sap/opu/odata/sap/ path plus POST /BuildingGetDetail. These are documented addresses, not verified operational URLs.'},
        {'id': 'inconsistent-access-description',
         'detail': 'Access prose specifies natural persons; technical context specifies a technical user. Authorization objects remain unspecified (offen).'},
        {'id': 'inconsistent-operation-description',
         'detail': 'Read-only operation is labelled NotIdempotence; restrictions mention unrelated Sell-from-Stock scenarios. The response section also calls the building a business entity.'},
        {'id': 'missing-contract',
         'detail': 'The saved files contain neither the WSDL nor ApiSchema.yaml. Required flags, cardinalities, decimal scale and physical-table mappings cannot be established from the HTML tables.'},
    ]
    report = {'provenance': provenance, 'documentation': DOCUMENTATION, 'apiMetadata': metadata,
              'model': model, 'issues': issues, 'structures': structures,
              'summary': {'structures': len(structures),
                          'fields': sum(len(structure['fields']) for structure in structures),
                          'includes': sum(len(structure['includes']) for structure in structures),
                          'verifiedPhysicalTableMappings': 0}}
    api = {
        'identifier': API_ID,
        'name': 'SAP RE-FX – Gebäudestammdaten lesen',
        'description': 'Lesender, synchroner SOAP-Service ZAPI_X4AI_BAPI_RE_BU_GET_DET für die Detaildaten eines Gebäudes aus SAP RE-FX. Die Dokumentation beschreibt 25 Service-Strukturen, darunter Gebäudestammdaten, Adresse, Bemessungen, Partner, Status und kundeneigene Erweiterungen. Kein Massenabruf. Die Katalogbeschreibung bleibt wegen widersprüchlicher Angaben und fehlender Tabellenzuordnung im Entwurf; der Service ist in der Quelle als ACTIVE dokumentiert.',
        'status': 'Entwurf', 'sourceStatus': metadata['apiState'],
        'version': metadata['apiVersion'], 'created': args.captured, 'modified': args.captured,
        'responsibleOrg': metadata['businessOwner'],
        'source': 'SAP API Dokumentation Bund',
        'sourceDetail': 'Building Master Data - Get Detail · Confluence-Seite 1105159761 · Version 7 · API-Stand ' + metadata['lastModification'] + ' · Zuordnung zum reduzierten Systemmodell ausstehend',
        'sourceUrl': DOCUMENTATION, 'synced': args.captured, 'provenance': provenance,
        'domain': 'bau', 'system': 'sap', 'protocol': metadata['apiTechnology'],
        'technicalName': metadata['apiName'], 'documentation': DOCUMENTATION,
        'accessRights': 'Autorisierung durch das Fachamt; für Drittanwendungen Vereinbarung mit dem BIT. Konkrete Berechtigungsobjekte nicht dokumentiert.',
        'operation': metadata['apiName'], 'httpMethod': 'POST', 'readOnly': True,
        'callType': 'synchronous', 'bulkSupported': False,
        'sourceModified': metadata['lastModification'],
        'endpointURL': servers[0]['url'],
        'endpointVerification': 'unverified', 'documentedServers': servers,
        'documentedPath': '/BuildingGetDetail',
        'wsdlDocumentation': 'https://confluence.bit.admin.ch/download/attachments/1105159761/ZAPI_X4AI_BAPI_RE_BU_GET_DET.wsdl',
        'authentication': [
            {'method': 'Basic Authentication', 'note': 'SAP-Benutzer über HTTPS'},
            {'method': 'SAP Logon Ticket', 'note': 'Vertrauensbasierte Authentifizierung (SSO)'},
            {'method': 'SAML / X.509', 'note': 'Optional über SAP PI/PO oder Middleware'},
        ],
        'oauthSupported': False,
        'documentedRequestParameters': [{'technicalName': 'BUILDING_ID', 'labels': {'de': 'Gebäude-Identifikator'},
                                         'dataType': 'String', 'mandatory': True,
                                         'verification': 'conflicts-with-structure-documentation'}],
        'responseStructures': [{'technicalName': structure['technicalName'], 'name': structure['name'],
                                'description': structure['description'], 'fieldCount': len(structure['fields']),
                                'sourceUrl': structure['sourceUrl']} for structure in structures],
        'errorHandling': 'Rückgabemeldungen im Service-Knoten RETURN auswerten, auch bei technisch erfolgreichem Aufruf. Laut Dokumentation höchstens 20 Meldungen; die letzte weist gegebenenfalls auf weitere Meldungen hin.',
        'documentationIssues': issues,
    }
    updated_tables = None
    related_updates = []
    if has_diagram:
        report_path = ROOT / 'docs/sap-refx-import-report.json'
        previous_report = json.loads(report_path.read_text(encoding='utf-8')) if report_path.exists() else {}
        updated_tables, reconciliation = reconcile_model(model, structures, provenance, args.captured)
        replacements = {item['table']: item for item in previous_report.get('reconciliation', {}).get('prototypeReplacements', [])}
        replacements.update({item['table']: item for item in reconciliation['prototypeReplacements']})
        reconciliation['prototypeReplacements'] = list(replacements.values())
        report['reconciliation'] = reconciliation
        updated_tables, curation = curate_catalog(updated_tables, model, structures, provenance, args.captured)
        related_updates = retire_references(curation, previous_report)
        report['catalogCuration'] = curation
        if API_ID in curation['decisions'].get('apiComments', {}):
            api['comment'] = curation['decisions']['apiComments'][API_ID]
        api['modelMappings'] = [item for item in reconciliation['candidates'] if item['table'] in curation['activeTableIds']]
        api['documentedFieldMappings'] = [
            {'table': table['identifier'], 'fieldId': field['technicalName'], 'structure': mapping['structure'],
             'field': mapping['field'], 'verification': mapping['verification'], 'physicalColumnVerified': False}
            for table in updated_tables if table.get('fieldScope') == 'api-projection'
            for field in table['fields'] for mapping in field.get('apiMappings', []) if mapping['api'] == API_ID]
        api['modelCoverage'] = {'scope': 'active-catalog', 'classes': curation['activeTables'],
                                'modelFields': curation['activeFields'],
                                'candidateModelFields': len(api['modelMappings']),
                                'documentedApiFields': len(api['documentedFieldMappings']), 'verifiedPhysicalTableMappings': 0}
        api['sourceReconciliation'] = {'scope': 'source-diagram-only', **reconciliation['summary'],
                                       'report': 'docs/sap-refx-import-report.json',
                                       'note': 'Source matches are retained as evidence. Removed Nutzungssicht entries are not active catalog mappings or approved requirements.'}
        api['sourceDetail'] = api['sourceDetail'].replace('Zuordnung zum reduzierten Systemmodell ausstehend', 'BUILDING als API-Feldprojektion dokumentiert; physische Spaltenzuordnung ausstehend')
        api['description'] = api['description'].replace('fehlender Tabellenzuordnung', 'nicht verifizierter technischer Feldzuordnung')
    apis_path = ROOT / 'data/apis.json'
    apis = json.loads(apis_path.read_text(encoding='utf-8'))
    existing = next((item for item in apis if item['identifier'] == API_ID), None)
    if existing:
        if existing.get('provenance', {}).get('importId') != IMPORT_ID:
            raise ValueError('API identifier is owned by another record')
        if existing.get('modelCoverage') and not has_diagram:
            raise ValueError('A diagram import exists; the empty workbook cannot replace its reconciliation')
        api['created'] = existing['created']
        preserve_annotations(api, existing)
        apis[apis.index(existing)] = api
    else:
        apis.append(api)
    write_json(ROOT / 'docs/sap-refx-import-report.json', report)
    write_json(apis_path, apis)
    if updated_tables is not None:
        write_json(ROOT / 'data/tables.json', updated_tables)
        for path, value in related_updates:
            write_json(path, value)
    print(json.dumps(report['summary']))
    if has_diagram:
        print(json.dumps(report['reconciliation']['summary']))
        print(json.dumps({key: report['catalogCuration'][key] for key in ['activeTables', 'activeFields']}))
    else:
        print('API imported; tables.json unchanged: the model workbook has no attribute records.')


if __name__ == '__main__':
    main()
