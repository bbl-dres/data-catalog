"""Import the reviewed GIS IMMO workbook. Requires openpyxl; no runtime dependency.

Source formats and DB field names describe the workbook model, not a live schema.
"""
import argparse
from collections import Counter
from copy import deepcopy
from datetime import date
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
IMPORT_ID = 'gis-immo-workbook'
HEADERS = ['ID', 'Status', 'Anwendung', 'Geschäftsobjekt', 'Merkmal Gruppe', 'Merkmal Alias DE',
           'Merkmal Alias EN', 'Merkmal DB', 'Merkmal DB Len', 'Format', 'Herkunft', 'Beschreibung DE']
KEYS = ['sourceId', 'status', 'application', 'entity', 'group', 'labelDe', 'labelEn',
        'technicalName', 'fieldNameLengthFormula', 'format', 'origin', 'description']
ENTITIES = {
    'BBL Gebäude': ('t-geb-gis', 'Gebäude', 'bau', 'gebaeude',
                    'Gebäudedaten mit BBL-Stammdaten, Adresse, Koordinaten, Angaben der amtlichen Vermessung sowie Flächen- und Volumenkennzahlen.'),
    'BBL Gebäude (AO)': ('t-boden', 'Bodenabdeckung', 'bau', 'bodenbedeckung',
                         'Bodenabdeckung mit dem Objekttyp Gebäude: Gebäudegrundfläche als Polygon. Das dokumentierte Typinventar umfasst Adress- und Vermessungsangaben, Bemessungen, Energiedaten und Referenzen auf externe Geometrie.'),
    'BBL Grundstück': ('t-parzelle', 'Grundstück', 'bau', 'grundstueck',
                       'Grundstücksdaten mit BBL-Stammdaten, Lage, amtlicher Vermessung, Flächenkennzahlen und Referenzen auf externe Geometrie.'),
    'BBL Gebäudehülle (AO)': ('t-huelle', 'Gebäudehülle (AO)', 'bau', 'gebaeude',
                             'Modellierte Angaben zur Gebäudehülle mit Lage, Flächen- und Volumenkennzahlen sowie Architektur-Objekt-ID und Geometriequelle.'),
    'BBL Raum': ('t-gis-room', 'Raum', 'bau', 'raum',
                 'Raumdaten mit BBL-Zuordnung, Abmessungen, Klassifizierung, Belegung, Reinigung und Referenzen auf externe Geometrie.'),
    'BBL Bauprojekt': ('t-proj', 'Bauprojekt', 'projekt', 'bauprojekt',
                       'Bauprojektdaten mit BBL-Stammdaten, Lage sowie Angaben zu Auftraggeber, Bauwerk, Projektkosten und Terminen.'),
    'BBL Grünfläche': ('t-gis-green-area', 'Grünfläche', 'bau', None,
                       'Grünflächendaten mit Grundstücksbezug, Lage sowie Architektur-Objekt-ID, Objekttyp, Flächenwert und Geometriequelle.'),
}
FIELD_NOTES = {
    ('BBL Gebäude', 'bbl_hist'): 'Der technische Name bbl_hist ist in der Quelle zweimal mit verschiedenen Bezeichnungen vergeben. Beide Definitionen bleiben getrennt; der korrekte Spaltenname ist zu klären.',
    ('BBL Grundstück', 'larea_uuf'): 'Die Quelle bezeichnet larea_uuf als bearbeitete Umgebungsfläche (BUF). Kürzel und Bezeichnung sind zu prüfen; die Quellangaben wurden unverändert übernommen.',
    ('BBL Grundstück', 'larea_buf'): 'Die Quelle bezeichnet larea_buf als unbearbeitete Umgebungsfläche (UUF). Kürzel und Bezeichnung sind zu prüfen; die Quellangaben wurden unverändert übernommen.',
    ('BBL Gebäude (AO)', 'av_egrid'): 'Die Quelle verwendet die Bezeichnung AV AGRID für av_egrid. Die Bezeichnung wurde unverändert übernommen; Schreibweise ist zu prüfen.',
    ('BBL Gebäudehülle (AO)', 'av_egrid'): 'Die Quelle verwendet die Bezeichnung AV AGRID für av_egrid. Die Bezeichnung wurde unverändert übernommen; Schreibweise ist zu prüfen.',
}


def read_json(path):
    return json.loads(path.read_text(encoding='utf-8'))


def write_json(path, value):
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')


def extract(path):
    workbook = load_workbook(path, data_only=False)
    cached = load_workbook(path, data_only=True)
    try:
        if workbook.sheetnames != ['Attribute']:
            raise ValueError('Workbook sheets changed; review the import scope')
        sheet = workbook['Attribute']
        if [cell.value for cell in sheet[1]][:12] != HEADERS:
            raise ValueError('Workbook headers changed; review the column mapping')
        rows = []
        for cells in sheet.iter_rows(min_row=2):
            if not any(cell.value is not None for cell in cells):
                continue
            if any(cell.value is not None for cell in cells[12:]):
                raise ValueError('Unexpected data outside the documented columns')
            row = {'row': cells[0].row, **dict(zip(KEYS, [cell.value for cell in cells[:12]]))}
            for key in ('application', 'entity', 'group', 'labelDe', 'technicalName', 'format'):
                if not isinstance(row[key], str) or not row[key].strip():
                    raise ValueError(f"Row {row['row']}: missing {key}")
            if any(cell.data_type == 'f' for cell in cells[:8] + cells[9:12]):
                raise ValueError('Unexpected formula outside the field-name length column')
            if row['application'] != 'BBL GIS IMMO' or row['status'] not in ('LIVE', 'DEV', None):
                raise ValueError(f"Row {row['row']}: unexpected application or source status")
            if row['fieldNameLengthFormula'] != '=LEN(Table2[[#This Row],[Merkmal DB]])':
                raise ValueError('The DB Len column changed meaning; review before importing')
            row['fieldNameLengthCached'] = cached['Attribute'].cell(row['row'], 9).value
            rows.append(row)
        if {row['entity'] for row in rows} != set(ENTITIES):
            raise ValueError('Model groups changed; review the catalog mapping before importing')
        source_ids = [row['sourceId'] for row in rows if row['sourceId'] is not None]
        if len(source_ids) != len(set(source_ids)):
            raise ValueError('Duplicate source row ID')
        return rows, workbook.properties.modified.isoformat() if workbook.properties.modified else None
    finally:
        workbook.close()
        cached.close()


def build_field(row, duplicate):
    name = row['technicalName']
    field = {'technicalName': name, 'technicalNameKind': 'model-attribute',
             'labels': {'de': row['labelDe']}, 'description': row['description'] or '',
             'dataType': row['format'], 'dataTypeKind': 'model-type', 'keyRole': None}
    if duplicate:
        source_key = row['sourceId'] if row['sourceId'] is not None else f"row-{row['row']}"
        field['identifier'] = f'{name}-source-{source_key}'
    if row['labelEn']:
        field['labels']['en'] = row['labelEn']
    if row['status']:
        field['sourceStatus'] = row['status']
    if row['origin']:
        field['source'] = row['origin']
    note = FIELD_NOTES.get((row['entity'], name))
    if not row['status']:
        note = 'Der Status ist in der Arbeitsmappe nicht angegeben; LIVE oder DEV wurde nicht abgeleitet.'
    if note:
        field['comment'] = note
    field['catalogMetadata'] = {
        'sourceSheet': 'Attribute', 'sourceRow': row['row'], 'sourceId': row['sourceId'],
        'modelEntity': row['entity'], 'attributeGroup': row['group'], 'origin': row['origin'],
        'sourceStatus': row['status'], 'reportedFormat': row['format'],
        'fieldNameLengthFormula': row['fieldNameLengthFormula'], 'fieldNameLengthCached': row['fieldNameLengthCached'],
    }
    return field


def import_catalog(path, captured, root=ROOT):
    rows, source_modified = extract(path)
    tables_path, systems_path = root / 'data/tables.json', root / 'data/systems.json'
    history_path, report_path = root / 'data/changelog.json', root / 'docs/sources/gis-immo/gis-immo-import-report.json'
    tables, systems, history = read_json(tables_path), read_json(systems_path), read_json(history_path)
    old_report = read_json(report_path) if report_path.exists() else {}
    before = {table['identifier']: table for table in tables}
    system = next(item for item in systems if item['identifier'] == 'gis')
    imported_ids = {entry[0] for entry in ENTITIES.values()}
    allowed = imported_ids
    if any(table['system'] == 'gis' and table['identifier'] not in allowed for table in tables):
        raise ValueError('Unreviewed GIS table present; refusing to replace its scope')
    domains = {item['identifier'] for item in read_json(root / 'data/domains.json')}
    objects = {item['identifier'] for item in read_json(root / 'data/objects.json')}
    if any(domain not in domains or (realizes and realizes not in objects)
           for _, _, domain, realizes, _ in ENTITIES.values()):
        raise ValueError('Catalog mapping references an unknown domain or business object')
    archive = deepcopy(old_report.get('replacedCatalog', {}))
    if not archive:
        archive = {'system': deepcopy(system), 'tables': [deepcopy(table) for table in tables if table['system'] == 'gis'],
                   'history': [deepcopy(entry) for entry in history if entry['entity'] in {'systems:gis', *(f'tables:{id}' for id in allowed)}]}
    provenance = {'importId': IMPORT_ID, 'captured': captured, 'sourceFile': path.name,
                  'sourceSha256': hashlib.sha256(path.read_bytes()).hexdigest(), 'sourceSheet': 'Attribute'}
    imported, summary = [], []
    for model_entity, (identifier, name, domain, realizes, description) in ENTITIES.items():
        previous = before.get(identifier)
        if previous and previous.get('system') != 'gis':
            raise ValueError(f'{identifier}: table identifier belongs to another system')
        members = [row for row in rows if row['entity'] == model_entity]
        duplicates = Counter(row['technicalName'] for row in members)
        old_fields = {field.get('identifier', field['technicalName']): field for field in (previous or {}).get('fields', [])}
        fields = []
        for row in members:
            duplicate = duplicates[row['technicalName']] > 1
            field = build_field(row, duplicate)
            old = old_fields.get(field.get('identifier', field['technicalName']))
            if old and 'comment' in old:
                field['comment'] = old['comment']
            fields.append(field)
        statuses = Counter(row['status'] or 'unspecified' for row in members)
        status_text = ', '.join(f'{count} {status}' for status, count in statuses.items())
        comment = f'Modellinventar laut Arbeitsmappe: {status_text}. Physischer Tabellenname, Schlüsselfelder und Spaltenlängen sind nicht dokumentiert. Die Beschreibung fasst die Feldgruppen zusammen.'
        if any(count > 1 for count in duplicates.values()):
            comment += ' bbl_hist ist zweimal unterschiedlich bezeichnet; beide Quellzeilen sind mit getrennten Katalog-IDs erhalten.'
        if model_entity == 'BBL Gebäude (AO)':
            comment += ' Fachliche Einordnung laut Katalogverantwortlichem: Typ Gebäude der Bodenabdeckung, als Gebäudegrundfläche (Polygon). Die Arbeitsmappe nennt diesen Typ BBL Gebäude (AO). Die Felder gelten für diesen Typ; eine gemeinsame Felddefinition für weitere Typen ist nicht belegt.'
        record = {'identifier': identifier, 'name': name, 'labels': {'de': name}, 'description': description,
                  'status': 'Entwurf', 'created': previous['created'] if previous else captured, 'modified': captured,
                  'system': 'gis', 'domain': domain, 'source': 'BBL GIS IMMO – Modellbeschreibung',
                  'sourceDetail': f'{path.name} · Attribute · {model_entity} · {len(fields)} Felddefinitionen',
                  'sourceModified': source_modified, 'synced': captured, 'modelClass': model_entity,
                  'fieldScope': 'model-inventory', 'sourceStatusCounts': dict(statuses),
                  'fields': fields, 'provenance': {**provenance, 'sourceRows': [row['row'] for row in members]},
                  'comment': (previous or {}).get('comment', comment)}
        if realizes:
            record['realizes'] = realizes
        if model_entity == 'BBL Gebäude (AO)':
            record['objectTypes'] = [{'name': 'Gebäude', 'businessObject': 'gebaeude', 'sourceClass': model_entity,
                                      'geometryType': 'Polygon', 'description': 'Gebäudegrundfläche als Polygon; fachliche Einordnung durch den Katalogverantwortlichen.',
                                      'fieldIds': [field.get('identifier', field['technicalName']) for field in fields]}]
            for field in fields:
                field['appliesToObjectTypes'] = ['Gebäude']
        for key in ('responsibleOrg', 'dataOwner', 'dataSteward', 'dataCustodian', 'classification', 'contact'):
            value = (previous or {}).get(key, system.get(key))
            if value is not None:
                record[key] = value
        if previous and previous.get('informationUrls'):
            record['informationUrls'] = previous['informationUrls']
        imported.append(record)
        summary.append({'tableId': identifier, 'modelEntity': model_entity, 'fieldCount': len(fields), 'sourceStatuses': dict(statuses)})
    by_id = {table['identifier']: table for table in imported}
    updated = [by_id.get(table['identifier'], table) for table in tables]
    updated.extend(table for table in imported if table['identifier'] not in before)
    if len({table['identifier'] for table in updated}) != len(updated):
        raise ValueError('Duplicate table identifier')
    for table in imported:
        ids = [field.get('identifier', field['technicalName']) for field in table['fields']]
        if len(ids) != len(set(ids)):
            raise ValueError('Duplicate catalog field identifier')
    system.update(description='Geoinformationssystem für Bundesimmobilien. Das dokumentierte Modell umfasst Gebäude, Bodenabdeckung mit dem Typ Gebäude, Grundstück, Gebäudehülle (AO), Raum, Bauprojekt und Grünfläche.',
                  modified=captured, source='BBL GIS IMMO – Modellbeschreibung', sourceDetail=f'{path.name} · Attribute · 7 Modellgruppen',
                  sourceModified=source_modified, synced=captured, provenance=provenance)
    system.pop('version', None)
    system.pop('personalData', None)
    system.setdefault('comment', 'Die Arbeitsmappe enthält Felddefinitionen mit den Quellenstatus LIVE und DEV. Der Katalogstatus bleibt Entwurf. Technische Tabellennamen, Geometrietypen und API-Verfügbarkeit sind daraus nicht bestätigt; vorhandene Zuständigkeiten bleiben Katalogangaben.')
    affected = {'systems:gis', *(f'tables:{id}' for id in allowed)}
    history = [entry for entry in history if entry['entity'] not in affected] if not old_report else [
        entry for entry in history if not (entry.get('importId') == IMPORT_ID and entry['date'] == captured)]
    for key in ['systems:gis', *(f'tables:{table["identifier"]}' for table in imported)]:
        history.append({'entity': key, 'date': captured, 'action': 'Modell importiert',
                        'detail': f'Felddefinitionen aus {path.name} übernommen; Quellstatus und offene Modellfragen separat dokumentiert.', 'user': 'Katalogimport',
                        'importId': IMPORT_ID})
    missing = {key: [row['row'] for row in rows if row[key] is None] for key in ('sourceId', 'status', 'labelEn', 'origin', 'description')}
    report = {'provenance': provenance, 'sourceModified': source_modified, 'columnMapping': dict(zip(HEADERS, KEYS)),
              'summary': {'tables': len(imported), 'fieldRows': len(rows), 'sourceStatuses': dict(Counter(row['status'] or 'unspecified' for row in rows))},
              'entities': summary, 'missingValues': missing,
              'duplicateNames': [{'tableId': table['identifier'], 'technicalName': name, 'rows': [field['catalogMetadata']['sourceRow'] for field in table['fields'] if field['technicalName'] == name]}
                                 for table in imported for name, count in Counter(field['technicalName'] for field in table['fields']).items() if count > 1],
              'retiredTableIds': [], 'sourceRows': rows, 'replacedCatalog': archive,
              'notes': ['DB Len is a LEN formula for identifier characters, not column storage length.',
                        'No physical table IDs, keys, nullability, geometry schema or code-list definitions are provided.',
                        'Reported formats, labels and descriptions are preserved without correcting source inconsistencies.',
                        'LIVE/DEV are workbook field statuses; they do not change the catalog Entwurf policy.',
                        'Owner classified the source group BBL Gebäude (AO) as type Gebäude of Bodenabdeckung, representing Gebäudegrundfläche as a polygon; t-boden retains type-specific fields.',
                        'Business-object definitions and API claims are unchanged. No classification of Grünfläche as another Bodenabdeckung type is inferred.']}
    for target, value in ((tables_path, updated), (systems_path, systems), (history_path, history), (report_path, report)):
        write_json(target, value)
    return report


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--workbook', type=Path, required=True)
    parser.add_argument('--captured', required=True)
    args = parser.parse_args()
    date.fromisoformat(args.captured)
    print(json.dumps(import_catalog(args.workbook, args.captured)['summary']))


if __name__ == '__main__':
    main()
